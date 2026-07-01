#!/usr/bin/env python3
"""
build_invoice_xlsx.py — Build the master 51-tab forensic invoice spreadsheet.

Usage:
    python build_invoice_xlsx.py \
        --ada forensic-output/_intermediate/ada_invoices.json \
        --devpipe forensic-output/_intermediate/devpipe_invoices.json \
        --extrapolated forensic-output/_intermediate/extrapolated.json \
        --commits forensic-output/_intermediate/commits.json \
        --git-by-month forensic-output/_intermediate/git_by_month.json \
        --project-name "Example Booking Co." \
        --out forensic-output/invoices/ExampleBooking_Invoice_Forensics.xlsx

The --commits and --git-by-month args are OPTIONAL — if not provided, the git-related
worksheets are skipped (relevant when no repository is available).
"""
import argparse, json, re, os
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from collections import defaultdict

# Sanitize openpyxl illegal chars
ILLEGAL_CHAR_RE = re.compile(r"[\x00-\x08\x0b\x0c\x0e-\x1f\x7f]")
import openpyxl.cell.cell as _ozc
_orig = _ozc.Cell._bind_value
def _safe_bind(self, v):
    if isinstance(v, str):
        v = ILLEGAL_CHAR_RE.sub("", v).strip()
    return _orig(self, v)
_ozc.Cell._bind_value = _safe_bind

HEADER_FILL = PatternFill(start_color='1F4E78', end_color='1F4E78', fill_type='solid')
HEADER_FONT = Font(bold=True, color='FFFFFF', size=11)
RECEIPT_FILL = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
EXTRAP_FILL = PatternFill(start_color='FFE699', end_color='FFE699', fill_type='solid')
THIN = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
TOTAL_FILL = PatternFill(start_color='FFEB9C', end_color='FFEB9C', fill_type='solid')
IDLE_FILL = PatternFill(start_color='FCE4E4', end_color='FCE4E4', fill_type='solid')

def style_header(ws, row, cols):
    for c in range(1, cols+1):
        cell = ws.cell(row=row, column=c)
        cell.fill = HEADER_FILL; cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = THIN

def autofit(ws, mw=80):
    for col_idx in range(1, ws.max_column+1):
        letter = ws.cell(row=1, column=col_idx).column_letter
        try:
            length = max(len(str(ws.cell(row=r, column=col_idx).value or '')) for r in range(1, ws.max_row+1))
        except Exception: length = 15
        ws.column_dimensions[letter].width = min(max(length+2, 10), mw)

def is_recurring_item(desc):
    d = (desc or '').lower()
    if re.search(r'\(\d{1,2}/\d{1,2}/\d{4}\s*-\s*\d{1,2}/\d{1,2}/\d{4}\)', desc or ''): return True
    return bool(re.search(r'\b(monthly|month|maintenance|hosting|workspace|gsuite|assistant|social media management|silver|subscription|platinum)\b', d))

def classify(items):
    if not items: return 'unknown'
    rec = sum(1 for it in items if is_recurring_item(it.get('description','')))
    if rec == len(items): return 'Recurring'
    if rec == 0: return 'One-off'
    return 'Mixed'

def main():
    ap = argparse.ArgumentParser()
    ap.add_argument('--ada', required=True)
    ap.add_argument('--devpipe', required=True)
    ap.add_argument('--extrapolated', required=True)
    ap.add_argument('--commits')
    ap.add_argument('--git-by-month')
    ap.add_argument('--project-name', required=True)
    ap.add_argument('--out', required=True)
    args = ap.parse_args()
    
    ada = json.load(open(args.ada))
    val = json.load(open(args.devpipe))
    extrap = json.load(open(args.extrapolated))
    commits = json.load(open(args.commits)) if args.commits and os.path.exists(args.commits) else []
    gbm = json.load(open(args.git_by_month)) if args.git_by_month and os.path.exists(args.git_by_month) else {}
    
    records = []
    for inv in ada:
        records.append({
            'invoice_num': inv['invoice_number'], 'date': inv['issued'] or inv['date_received'],
            'from': inv['vendor'], 'amount': inv['amount'], 'taxes': 0.0,
            'items': inv['items'], 'is_receipt': False,
            'pdf_or_source': f"Email (M-ref: {inv.get('message_id','')})",
            'is_extrapolated': False,
        })
    for inv in val:
        if inv['is_receipt']:
            date = inv['paid_date'] or inv['issued']
            amount = inv['total']
            invnum = (inv['receipt_number'] or '') + (f" (receipt for {inv['invoice_number']})" if inv['receipt_number'] else inv.get('invoice_number',''))
        else:
            date = inv['issued']; amount = inv['total']; invnum = inv['invoice_number']
        records.append({
            'invoice_num': invnum, 'date': date, 'from': inv['vendor'], 'amount': amount, 'taxes': 0.0,
            'items': inv['items'], 'is_receipt': inv['is_receipt'], 'pdf_or_source': inv['pdf_file'],
            'is_extrapolated': False,
        })
    for e in extrap:
        records.append({
            'invoice_num': e['invoice'], 'date': e['date'], 'from': e['vendor'],
            'amount': e['amount'], 'taxes': 0.0,
            'items': [{'description': e['item'], 'qty': e.get('qty', 1), 'unit_price': e['amount'], 'amount': e['amount']}],
            'is_receipt': False, 'pdf_or_source': 'EXTRAPOLATED (' + e['reason'][:80] + ')',
            'is_extrapolated': True, 'reason': e['reason'],
        })
    records.sort(key=lambda r: r['date'] or '9999')
    for r in records: r['classification'] = classify(r['items'])
    
    wb = Workbook()
    ws = wb.active; ws.title = 'Invoice Summary'
    headers = ['Invoice #', 'Date', 'From / Vendor', 'Amount', 'Taxes', 'Class', 'Source', '# Items', 'Extrapolated?']
    for c, h in enumerate(headers, 1): ws.cell(row=1, column=c, value=h)
    style_header(ws, 1, len(headers))
    
    doc_ada = doc_val = ext_ada = ext_val = 0
    for i, rec in enumerate(records, 2):
        ws.cell(row=i, column=1, value=rec['invoice_num'])
        ws.cell(row=i, column=2, value=rec['date'])
        ws.cell(row=i, column=3, value=rec['from'])
        ws.cell(row=i, column=4, value=rec['amount']).number_format = '"$"#,##0.00'
        ws.cell(row=i, column=5, value=rec['taxes']).number_format = '"$"#,##0.00'
        ws.cell(row=i, column=6, value=('Receipt' if rec['is_receipt'] else rec['classification']))
        ws.cell(row=i, column=7, value=rec['pdf_or_source'])
        ws.cell(row=i, column=8, value=len(rec['items']))
        ws.cell(row=i, column=9, value='YES' if rec['is_extrapolated'] else 'No')
        if rec['is_extrapolated']:
            for c in range(1, 10): ws.cell(row=i, column=c).fill = EXTRAP_FILL
        elif rec['is_receipt']:
            for c in range(1, 10): ws.cell(row=i, column=c).fill = RECEIPT_FILL
        for c in range(1, 10): ws.cell(row=i, column=c).border = THIN
        if not rec['is_receipt']:
            amt = rec['amount'] or 0
            is_ada = 'ADA' in rec['from']
            if rec['is_extrapolated']:
                if is_ada: ext_ada += amt
                else: ext_val += amt
            else:
                if is_ada: doc_ada += amt
                else: doc_val += amt
    
    trow = len(records) + 3
    ws.cell(row=trow, column=3, value='ADA Documented:').font = Font(bold=True)
    ws.cell(row=trow, column=4, value=doc_ada).number_format = '"$"#,##0.00'
    ws.cell(row=trow+1, column=3, value='ADA Extrapolated:').font = Font(bold=True)
    ws.cell(row=trow+1, column=4, value=ext_ada).number_format = '"$"#,##0.00'
    ws.cell(row=trow+2, column=3, value='DevPipe/Offshore Build Documented:').font = Font(bold=True)
    ws.cell(row=trow+2, column=4, value=doc_val).number_format = '"$"#,##0.00'
    ws.cell(row=trow+3, column=3, value='DevPipe/Offshore Build Extrapolated:').font = Font(bold=True)
    ws.cell(row=trow+3, column=4, value=ext_val).number_format = '"$"#,##0.00'
    grand = doc_ada + doc_val + ext_ada + ext_val
    ws.cell(row=trow+5, column=3, value='GRAND TOTAL:').font = Font(bold=True, size=12)
    ws.cell(row=trow+5, column=4, value=grand).number_format = '"$"#,##0.00'
    ws.cell(row=trow+5, column=4).font = Font(bold=True, size=12)
    ws.cell(row=trow+5, column=4).fill = TOTAL_FILL
    
    autofit(ws)
    
    # Git Commit Log if available
    if commits:
        ws_g = wb.create_sheet('Commit Log')
        h = ['Hash', 'Date', 'Author', 'Email', 'Subject', 'Files', 'LOC +', 'LOC -', 'Hours', '$ @ $100/hr']
        for c, name in enumerate(h, 1): ws_g.cell(row=1, column=c, value=name)
        style_header(ws_g, 1, len(h))
        for i, c in enumerate(commits, 2):
            ws_g.cell(row=i, column=1, value=c['hash'][:8])
            ws_g.cell(row=i, column=2, value=c['date'][:10])
            ws_g.cell(row=i, column=3, value=c['author'])
            ws_g.cell(row=i, column=4, value=c['email'])
            ws_g.cell(row=i, column=5, value=c['subject'][:200])
            ws_g.cell(row=i, column=6, value=c['files_changed'])
            ws_g.cell(row=i, column=7, value=c['insertions'])
            ws_g.cell(row=i, column=8, value=c['deletions'])
            ws_g.cell(row=i, column=9, value=c['est_hours'])
            ws_g.cell(row=i, column=10, value=c['est_hours'] * 100).number_format = '"$"#,##0.00'
        autofit(ws_g)
        
        ws_m = wb.create_sheet('Monthly Effort Rollup')
        h2 = ['Month', 'Commits', 'LOC +', 'LOC -', 'Net LOC', 'Hours', '$ @ $100/hr', 'Status']
        for c, name in enumerate(h2, 1): ws_m.cell(row=1, column=c, value=name)
        style_header(ws_m, 1, len(h2))
        for i, ym in enumerate(sorted(gbm.keys()), 2):
            d = gbm[ym]
            ws_m.cell(row=i, column=1, value=ym)
            ws_m.cell(row=i, column=2, value=d['commits'])
            ws_m.cell(row=i, column=3, value=d['insertions'])
            ws_m.cell(row=i, column=4, value=d['deletions'])
            ws_m.cell(row=i, column=5, value=d['insertions'] - d['deletions'])
            ws_m.cell(row=i, column=6, value=round(d['hours'], 1))
            ws_m.cell(row=i, column=7, value=d['hours'] * 100).number_format = '"$"#,##0.00'
            status = 'IDLE' if d['commits'] == 0 else ('Low' if d['hours'] < 10 else ('Heavy' if d['hours'] > 30 else 'Normal'))
            ws_m.cell(row=i, column=8, value=status)
            if d['commits'] == 0:
                for c in range(1, 9): ws_m.cell(row=i, column=c).fill = IDLE_FILL
        autofit(ws_m)
    
    # Per-invoice line item sheets (documented only; skip extrapolated)
    for rec in records:
        if rec['is_extrapolated']: continue
        sheet_name = re.sub(r'[\[\]\:\*\?\/\\]', '_', (rec['invoice_num'] or 'unknown'))[:25]
        suffix = '_R' if rec['is_receipt'] else ''
        sheet_name = sheet_name + suffix
        base = sheet_name; n = 1
        while sheet_name in wb.sheetnames:
            sheet_name = base + f'_{n}'; n += 1
        ws2 = wb.create_sheet(title=sheet_name[:31])
        ws2['A1'] = 'Invoice #:'; ws2['B1'] = rec['invoice_num']
        ws2['A2'] = 'Date:'; ws2['B2'] = rec['date']
        ws2['A3'] = 'Vendor:'; ws2['B3'] = rec['from']
        ws2['A5'] = 'Total:'; ws2['B5'] = rec['amount']
        ws2['B5'].number_format = '"$"#,##0.00'
        for r in range(1, 7): ws2.cell(row=r, column=1).font = Font(bold=True)
        h3 = ['Description', 'Date', 'Qty', 'Unit', 'Total', 'Vendor']
        for c, n in enumerate(h3, 1): ws2.cell(row=8, column=c, value=n)
        style_header(ws2, 8, len(h3))
        for i, it in enumerate(rec['items'], 9):
            ws2.cell(row=i, column=1, value=it.get('description', ''))
            ws2.cell(row=i, column=2, value=rec['date'])
            ws2.cell(row=i, column=3, value=it.get('qty', 1))
            ws2.cell(row=i, column=4, value=it.get('unit_price', it.get('amount', 0))).number_format = '"$"#,##0.00'
            ws2.cell(row=i, column=5, value=it.get('amount', 0)).number_format = '"$"#,##0.00'
            ws2.cell(row=i, column=6, value=rec['from'])
        autofit(ws2)
    
    os.makedirs(os.path.dirname(args.out), exist_ok=True)
    wb.save(args.out)
    print(f"Wrote {args.out} ({len(wb.sheetnames)} sheets)")
    print(f"ADA documented: ${doc_ada:,.2f}, extrapolated: ${ext_ada:,.2f}")
    print(f"DevPipe documented: ${doc_val:,.2f}, extrapolated: ${ext_val:,.2f}")
    print(f"GRAND TOTAL: ${grand:,.2f}")

if __name__ == '__main__':
    main()
